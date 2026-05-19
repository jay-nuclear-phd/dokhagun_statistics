import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime, timedelta
import os

# Configuration
CSV_PATH = 'data/result.csv'
TEMPLATE_PATH = 'data/독서모임 글쓰기 진도.xlsx'
START_DATE = datetime(2026, 5, 4)  # 2026.05.04 is a Monday
TODAY = datetime(2026, 5, 13)

def get_week_number(date_obj):
    delta = (date_obj - START_DATE).days
    if delta < 0:
        return -1 # Before start
    return (delta // 7) + 1

def generate_report():
    # 1. Read CSV
    df_csv = pd.read_csv(CSV_PATH)
    # Parse date: 2026.05.10. -> 2026-05-10
    df_csv['date_dt'] = pd.to_datetime(df_csv['date'].str.rstrip('.'), format='%Y.%m.%d')
    df_csv['week'] = df_csv['date_dt'].apply(get_week_number)
    
    # 2. Extract participants from template
    wb_temp = openpyxl.load_workbook(TEMPLATE_PATH)
    ws_temp = wb_temp.active
    
    template_participants = []
    for col_idx in range(2, ws_temp.max_column + 1):
        nick = ws_temp.cell(row=2, column=col_idx).value
        if nick:
            template_participants.append(str(nick))
            
    # 3. Group by writer and week, count posts
    counts = df_csv.groupby(['writer', 'week']).size().reset_index(name='post_count')
    
    # Helper to normalize names for comparison
    def normalize(name):
        return str(name).replace(" ", "").lower()

    # Identify all unique writers and map them to template participants or new ones
    all_participants = list(template_participants)
    writer_to_nick = {}
    
    for _, row in counts.iterrows():
        writer = str(row['writer'])
        norm_writer = normalize(writer)
        
        found = False
        for nick in template_participants:
            norm_nick = normalize(nick)
            if norm_nick in norm_writer or norm_writer in norm_nick:
                writer_to_nick[writer] = nick
                found = True
                break
        
        if not found:
            if writer not in all_participants:
                all_participants.append(writer)
            writer_to_nick[writer] = writer

    # Create mapping: (nickname, week) -> count
    count_map = {}
    for _, row in counts.iterrows():
        nick = writer_to_nick[row['writer']]
        week = row['week']
        key = (nick, week)
        count_map[key] = count_map.get(key, 0) + row['post_count']
        
    # 4. Load styles
    status_styles = {}
    found = set()
    for row in ws_temp.iter_rows():
        for cell in row:
            val = str(cell.value).upper() if cell.value else ''
            if val == 'O' and 'O' not in found:
                status_styles['O'] = (cell.value, cell.fill, cell.font, cell.border, cell.alignment)
                found.add('O')
            elif val == 'X' and 'X' not in found:
                status_styles['X'] = (cell.value, cell.fill, cell.font, cell.border, cell.alignment)
                found.add('X')
            elif val == '☆' and '☆' not in found:
                status_styles['☆'] = (cell.value, cell.fill, cell.font, cell.border, cell.alignment)
                found.add('☆')
        if len(found) == 3: break
    
    if 'O' not in status_styles: status_styles['O'] = ('O', None, None, None, None)
    if 'X' not in status_styles: status_styles['X'] = ('X', None, None, None, None)
    if '☆' not in status_styles: status_styles['☆'] = ('☆', None, None, None, None)

    # 5. Create New Workbook
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    ws_new.title = "진도표"
    
    headers = [''] + all_participants
    ws_new.append(headers)
    
    rows_labels = []
    for w in range(1, 5):
        for n in range(1, 4):
            rows_labels.append(f"{w}주차 {n}번")
            
    current_week = get_week_number(TODAY)
    
    for label in rows_labels:
        week_num = int(label.split('주차')[0])
        post_num = int(label.split(' ')[1].replace('번', ''))
        
        row_data = [label]
        for nick in all_participants:
            count = count_map.get((nick, week_num), 0)
            
            if count >= post_num:
                status = 'O'
            else:
                if week_num < current_week:
                    status = 'X'
                elif week_num == current_week:
                    status = '☆'
                else:
                    status = '' # Future weeks blank
            
            row_data.append(status)
        
        ws_new.append(row_data)
        
        curr_row = ws_new.max_row
        for col_idx, status in enumerate(row_data):
            cell = ws_new.cell(row=curr_row, column=col_idx + 1)
            
            if col_idx == 0:
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal='center')
                continue

            if status == '':
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                continue

            style_info = status_styles.get(status)
            if style_info:
                _, fill, font, border, alignment = style_info
                if fill: cell.fill = PatternFill(start_color=fill.start_color, end_color=fill.end_color, fill_type=fill.fill_type)
                if font: cell.font = Font(name=font.name, size=font.size, bold=font.bold, italic=font.italic, color=font.color)
                if border: cell.border = Border(left=border.left, right=border.right, top=border.top, bottom=border.bottom)
                if alignment: cell.alignment = Alignment(horizontal=alignment.horizontal, vertical=alignment.vertical)
            
            if not cell.border.left:
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for col_idx in range(1, len(headers) + 1):
        cell = ws_new.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        thin = Side(border_style="thin", color="000000")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    date_str = TODAY.strftime('%Y_%m_%d')
    output_filename = f'bookclub_writeup_counting_as_of_{date_str}.xlsx'
    output_path = os.path.join('data', output_filename)
    wb_new.save(output_path)
    print(f"File saved: {output_path}")

if __name__ == "__main__":
    generate_report()
